import io
from flask import Blueprint, render_template, abort, send_file, Response
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload

from app.models import Employee, TrainingSession, Location

legacy_bp = Blueprint('legacy', __name__, url_prefix='/legacy')


def _require_manager():
    if not current_user.is_authenticated or current_user.role != 'manager':
        abort(403)


def _build_trainee_data():
    trainees_raw = (
        Employee.query
        .filter_by(role='trainee', status='active')
        .filter(Employee.graduated_at == None)
        .options(
            joinedload(Employee.trainee_sessions)
            .joinedload(TrainingSession.ratings),
            joinedload(Employee.trainee_sessions)
            .joinedload(TrainingSession.position),
        )
        .order_by(Employee.first_name)
        .all()
    )

    trainees = []
    for emp in trainees_raw:
        pos_map = {}
        for sess in emp.trainee_sessions:
            if not sess.completed_at or not sess.position:
                continue
            pos_name = sess.position.name
            pos_map.setdefault(pos_name, []).append(sess)

        positions = []
        for pos_name, sessions in sorted(pos_map.items()):
            all_ratings = [r.rating_value for s in sessions for r in s.ratings]
            avg_rating = round(sum(all_ratings) / len(all_ratings), 1) if all_ratings else 0.0
            pct = round((avg_rating / 5) * 100, 1)
            positions.append({
                'name': pos_name,
                'avg_rating': avg_rating,
                'pct': pct,
                'session_count': len(sessions),
            })

        overall_pct = round(sum(p['pct'] for p in positions) / len(positions), 1) if positions else 0.0
        trainees.append({
            'id': emp.id,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'start_date': emp.start_date.strftime('%b %d, %Y') if emp.start_date else '—',
            'location': emp.location.name if emp.location else 'General',
            'positions': positions,
            'overall_pct': overall_pct,
        })

    # Group by location
    groups = {}
    for t in trainees:
        loc = t['location']
        groups.setdefault(loc, []).append(t)

    return trainees, groups


@legacy_bp.route('/')
@login_required
def index():
    _require_manager()
    trainees, groups = _build_trainee_data()
    return render_template('legacy.html', trainees=trainees, groups=groups)


@legacy_bp.route('/export')
@login_required
def export_excel():
    _require_manager()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response("openpyxl not installed.", status=500)

    trainees, _ = _build_trainee_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Progress"

    # Header style (matching original: blue #004F72 header with white text)
    hdr_fill = PatternFill("solid", fgColor="004F72")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Collect all positions across trainees
    all_positions = sorted({p['name'] for t in trainees for p in t['positions']})

    # Write header row
    headers = ['Name', 'Start Date', 'Location', 'Overall %'] + all_positions
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Color helpers (matching original progress colors)
    def progress_fill(pct):
        if pct >= 80:   return PatternFill("solid", fgColor="C6EFCE")  # green
        if pct >= 70:   return PatternFill("solid", fgColor="FFEB9C")  # yellow
        if pct >= 50:   return PatternFill("solid", fgColor="FFCC99")  # orange
        return              PatternFill("solid", fgColor="FFC7CE")      # red

    # Data rows
    pos_pct_map = {p: i for i, p in enumerate(all_positions)}
    for row_idx, t in enumerate(trainees, 2):
        pos_lookup = {p['name']: p['pct'] for p in t['positions']}

        ws.cell(row=row_idx, column=1, value=f"{t['first_name']} {t['last_name']}")
        ws.cell(row=row_idx, column=2, value=t['start_date'])
        ws.cell(row=row_idx, column=3, value=t['location'])

        overall_cell = ws.cell(row=row_idx, column=4, value=f"{t['overall_pct']}%")
        overall_cell.fill = progress_fill(t['overall_pct'])
        overall_cell.font = Font(bold=True)
        overall_cell.alignment = Alignment(horizontal="center")

        for pos_name in all_positions:
            col = 4 + pos_pct_map[pos_name] + 1
            pct = pos_lookup.get(pos_name, None)
            if pct is not None:
                c = ws.cell(row=row_idx, column=col, value=f"{pct}%")
                c.fill = progress_fill(pct)
                c.alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=row_idx, column=col, value="—")

        # Style base cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = cell_border

    # Auto-width columns
    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, len(trainees) + 2)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='training_progress_classic.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
